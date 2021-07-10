import sqlite3
from openpyxl import load_workbook

conn = sqlite3.connect('Movie.db')
wb = load_workbook('Movie.csv')
ws = wb['Movie']

conn.execute("create table if not exists Movies (movie text, seat text, price int)")

for i in range(1,10):
    temp_str = "insert into food_items (movie, seat, price) values ('{0}', '{1}', '{2}')".format(ws.cell(i,1).value, ws.cell(i,2).value, ws.cell(i,3).value)
    conn.execute(temp_str)

conn.commit()

content = conn.execute("select * from Movies")
for i in content:
    print(i)
